from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO
import json
from pathlib import Path
import re
import ssl
from typing import Any
import unicodedata
from urllib.parse import urlencode, urlparse
from urllib.request import Request, urlopen

import numpy as np
import pandas as pd
import streamlit as st
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps


BASE_DIR = Path(__file__).resolve().parent
PORTFOLIO_SAMPLE_PATH = BASE_DIR / "portfolio_sample.csv"
MARGIN_SAMPLE_PATH = BASE_DIR / "margin_sample.csv"
STOP_HISTORY_SAMPLE_PATH = BASE_DIR / "stop_history_sample.csv"

PORTFOLIO_REQUIRED_COLUMNS = ["symbol", "name", "shares", "cost", "category"]
MARGIN_REQUIRED_COLUMNS = [
    "symbol",
    "date",
    "margin_balance",
    "margin_change",
    "margin_change_rate",
    "foreign_buy_sell",
    "investment_trust_buy_sell",
    "price_change_rate",
]
STOP_HISTORY_REQUIRED_COLUMNS = ["symbol", "name", "last_trailing_stop", "last_update_date"]

DEFAULT_ATR_MULTIPLIERS = {
    "ETF": 2.0,
    "financial": 2.0,
    "low_volatility": 2.0,
    "normal": 2.0,
    "high_volatility": 2.5,
}

OCR_PORTFOLIO_COLUMNS = ["symbol", "name", "shares", "cost", "category", "image_name", "ocr_text"]
PORTFOLIO_IMAGE_TYPES = ["png", "jpg", "jpeg", "webp"]
OCR_HEADER_KEYWORDS = [
    "代號",
    "名稱",
    "股數",
    "庫存",
    "均價",
    "成本",
    "現價",
    "損益",
    "市值",
    "張數",
    "可賣",
]
OCR_NON_NAME_TOKENS = {
    "現股買進",
    "現股賣出",
    "零股買進",
    "零股賣出",
    "融資買進",
    "融資賣出",
    "融券賣出",
    "融券回補",
}

REPORT_COLUMNS = [
    "股票代號",
    "股票名稱",
    "股數",
    "成本價",
    "現價",
    "市值",
    "投入成本",
    "未實現損益",
    "未實現損益率",
    "最近高點",
    "最新TR",
    "ATR",
    "ATR倍數",
    "原始ATR移動停利價",
    "最終ATR移動停利價",
    "距離停利價差",
    "距離停利百分比",
    "風險狀態",
    "操作建議",
    "融資餘額",
    "融資增減",
    "融資增減率",
    "外資買賣超",
    "投信買賣超",
    "股價漲跌幅",
    "融資風險",
    "資料更新日期",
]

DISPLAY_PERCENT_COLUMNS = ["未實現損益率", "距離停利百分比", "融資增減率", "股價漲跌幅"]
DISPLAY_NUMBER_COLUMNS = [
    "成本價",
    "現價",
    "市值",
    "投入成本",
    "未實現損益",
    "最近高點",
    "最新TR",
    "ATR",
    "ATR倍數",
    "原始ATR移動停利價",
    "最終ATR移動停利價",
    "距離停利價差",
    "融資餘額",
    "融資增減",
    "外資買賣超",
    "投信買賣超",
]

RISK_DOWNLOAD_FAILURE = "資料下載失敗，暫不判斷"
RISK_DATA_INSUFFICIENT = "資料不足，暫不判斷"
OFFICIAL_MARGIN_LOOKBACK_DAYS = 10
OFFICIAL_DATA_TIMEOUT_SECONDS = 15
OFFICIAL_DATA_RETRY_ATTEMPTS = 2
OFFICIAL_MARGIN_CACHE_TTL_SECONDS = 300
OFFICIAL_DATA_USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36"
)


def format_ocr_error_message(error: Exception) -> str:
    """將常見 OCR 載入錯誤轉成較可操作的提示。"""
    message = str(error)
    lowered = message.lower()

    if "libgl.so.1" in lowered:
        return (
            "部署環境缺少 libGL 系統套件，導致 OCR 模組無法載入。"
            "若部署在 Streamlit Community Cloud，請確認專案根目錄有 packages.txt，"
            "並讓平台重新部署以安裝 libgl1 與對應 Debian 版本的 libglib runtime。"
        )

    if "libgthread-2.0.so.0" in lowered or "libglib-2.0.so.0" in lowered:
        return (
            "部署環境缺少 glib runtime，導致 OCR 模組無法載入。"
            "若部署在 Streamlit Community Cloud，請確認 packages.txt 使用與目前 Debian 版本相容的 glib 套件"
            "（例如 trixie 環境使用 libglib2.0-0t64），再重新部署。"
        )

    return message


def make_empty_margin_df() -> pd.DataFrame:
    """建立空白融資資料表，方便缺省情境直接沿用。"""
    return pd.DataFrame(columns=MARGIN_REQUIRED_COLUMNS)


def make_empty_stop_history_df() -> pd.DataFrame:
    """建立空白停利歷史資料表，供首次計算或缺檔時使用。"""
    return pd.DataFrame(columns=STOP_HISTORY_REQUIRED_COLUMNS)


def make_empty_ocr_portfolio_df() -> pd.DataFrame:
    """建立空白的 OCR 持股草稿資料表。"""
    return pd.DataFrame(columns=OCR_PORTFOLIO_COLUMNS)


def validate_required_columns(df: pd.DataFrame, required_columns: list[str]) -> list[str]:
    """檢查資料表是否缺少必要欄位。"""
    return [column for column in required_columns if column not in df.columns]


def normalize_symbol(symbol: Any) -> str:
    """標準化股票代號，若未帶市場尾碼則預設補上 .TW。"""
    cleaned = str(symbol).strip().upper()
    if cleaned and "." not in cleaned:
        cleaned = f"{cleaned}.TW"
    return cleaned


def get_symbol_code(symbol: Any) -> str:
    """取出股票代號主體，忽略 .TW / .TWO 市場尾碼。"""
    cleaned = str(symbol).strip().upper()
    if not cleaned:
        return ""
    return cleaned.split(".", 1)[0]


def build_symbol_lookup_candidates(symbol: Any) -> list[str]:
    """建立 Yahoo Finance 查價候選代號，支援上市與上櫃尾碼 fallback。"""
    cleaned = str(symbol).strip().upper()
    if not cleaned:
        return []

    if "." not in cleaned:
        return [f"{cleaned}.TW", f"{cleaned}.TWO"]

    symbol_code, market_suffix = cleaned.rsplit(".", 1)
    candidates = [cleaned]
    if market_suffix == "TW":
        candidates.append(f"{symbol_code}.TWO")
    elif market_suffix == "TWO":
        candidates.append(f"{symbol_code}.TW")
    return candidates


def to_float(value: Any) -> float:
    """將輸入轉成浮點數；若失敗則回傳 NaN。"""
    series = pd.to_numeric(pd.Series([value]), errors="coerce")
    result = series.iloc[0]
    return float(result) if pd.notna(result) else float("nan")


def safe_excel_value(value: Any) -> Any:
    """把 pandas 的遺漏值轉成 openpyxl 可接受的 None。"""
    return None if pd.isna(value) else value


def load_local_csv_bytes(file_path: Path) -> bytes:
    """讀取專案內建 CSV 並輸出成 UTF-8 BOM 下載內容。"""
    return pd.read_csv(file_path).to_csv(index=False).encode("utf-8-sig")


def parse_official_number(value: Any) -> float:
    """把官方 JSON 內帶千分位或百分號的數字字串轉成浮點數。"""
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    if not text:
        return float("nan")

    normalized = re.sub(r"\s+", "", text)
    normalized = normalized.replace(",", "").replace("%", "")
    normalized = normalized.replace("＋", "+").replace("－", "-")
    if normalized in {"", "-", "--", "---"}:
        return float("nan")

    result = pd.to_numeric(pd.Series([normalized]), errors="coerce").iloc[0]
    return float(result) if pd.notna(result) else float("nan")


def format_roc_date(target_date: date) -> str:
    """把西元日期轉成櫃買中心端點需要的民國日期字串。"""
    return f"{target_date.year - 1911}/{target_date.month:02d}/{target_date.day:02d}"


def build_official_ssl_context_candidates(url: str) -> list[ssl.SSLContext | None]:
    """為官方 JSON 端點建立可接受的 SSL context 候選清單。"""
    contexts: list[ssl.SSLContext | None] = [None]
    hostname = (urlparse(url).hostname or "").lower()

    # TPEX 偶爾會回傳讓部分 OpenSSL 組合驗證失敗的憑證鏈，
    # 這裡只對官方唯讀公開端點做有限 fallback，避免整批上櫃資料直接缺失。
    if hostname != "www.tpex.org.tw":
        return contexts

    try:
        import certifi

        contexts.append(ssl.create_default_context(cafile=certifi.where()))
    except Exception:
        pass

    contexts.append(ssl._create_unverified_context())
    return contexts


def fetch_url_bytes(url: str, request: Request) -> bytes:
    """下載官方端點內容，必要時重試並切換 SSL 驗證策略。"""
    recent_errors: list[str] = []

    for _ in range(OFFICIAL_DATA_RETRY_ATTEMPTS):
        for ssl_context in build_official_ssl_context_candidates(url):
            try:
                open_kwargs: dict[str, Any] = {"timeout": OFFICIAL_DATA_TIMEOUT_SECONDS}
                if ssl_context is not None:
                    open_kwargs["context"] = ssl_context
                with urlopen(request, **open_kwargs) as response:
                    return response.read()
            except Exception as exc:
                recent_errors.append(str(exc))

    summarized_errors: list[str] = []
    for message in recent_errors:
        if message not in summarized_errors:
            summarized_errors.append(message)

    error_text = "；".join(summarized_errors[-3:]) if summarized_errors else "未知錯誤"
    raise RuntimeError(f"下載官方資料失敗：{error_text}")


def fetch_json_from_url(url: str) -> dict[str, Any]:
    """以瀏覽器樣式 User-Agent 下載官方 JSON。"""
    request = Request(
        url,
        headers={
            "Accept": "application/json,text/plain,*/*",
            "User-Agent": OFFICIAL_DATA_USER_AGENT,
        },
    )
    return json.loads(fetch_url_bytes(url, request).decode("utf-8"))


def build_twse_margin_url(target_date: date) -> str:
    """建立證交所融資融券彙總 JSON 端點。"""
    query = urlencode(
        {
            "date": target_date.strftime("%Y%m%d"),
            "selectType": "ALL",
            "response": "json",
        }
    )
    return f"https://www.twse.com.tw/rwd/zh/marginTrading/MI_MARGN?{query}"


def build_tpex_margin_url(target_date: date) -> str:
    """建立櫃買中心融資融券餘額 JSON 端點。"""
    query = urlencode(
        {
            "l": "zh-tw",
            "o": "json",
            "d": format_roc_date(target_date),
            "s": "0,asc,0",
        }
    )
    return f"https://www.tpex.org.tw/web/stock/margin_trading/margin_balance/margin_bal_result.php?{query}"


def build_official_margin_row(
    symbol: str,
    record_date: str,
    current_balance: Any,
    previous_balance: Any,
) -> dict[str, Any]:
    """把官方欄位整理成系統既有的 margin.csv 結構。"""
    current_value = parse_official_number(current_balance)
    previous_value = parse_official_number(previous_balance)

    if pd.notna(current_value) and pd.notna(previous_value):
        margin_change = current_value - previous_value
    else:
        margin_change = float("nan")

    if pd.notna(margin_change) and pd.notna(previous_value) and previous_value != 0:
        margin_change_rate = (margin_change / previous_value) * 100
    else:
        margin_change_rate = float("nan")

    return {
        "symbol": symbol,
        "date": record_date,
        "margin_balance": current_value,
        "margin_change": margin_change,
        "margin_change_rate": margin_change_rate,
        "foreign_buy_sell": float("nan"),
        "investment_trust_buy_sell": float("nan"),
        "price_change_rate": float("nan"),
    }


def parse_twse_margin_payload(payload: dict[str, Any]) -> tuple[pd.DataFrame, str]:
    """解析證交所融資融券彙總資料。"""
    status_text = str(payload.get("stat", "")).strip()
    if status_text != "OK":
        raise ValueError(status_text or "證交所端點未回傳有效資料")

    record_date = pd.to_datetime(payload.get("date"), format="%Y%m%d", errors="coerce")
    if pd.isna(record_date):
        raise ValueError("證交所資料缺少有效日期")

    detail_table = None
    for table in payload.get("tables", []):
        fields = table.get("fields") or []
        if fields and fields[0] == "代號" and "前日餘額" in fields and "今日餘額" in fields:
            detail_table = table
            break

    if detail_table is None or not detail_table.get("data"):
        raise ValueError("證交所該日無融資融券明細")

    fields = detail_table.get("fields") or []
    previous_balance_index = fields.index("前日餘額")
    current_balance_index = fields.index("今日餘額")
    normalized_date = record_date.strftime("%Y-%m-%d")

    rows: list[dict[str, Any]] = []
    for item in detail_table.get("data", []):
        if len(item) <= current_balance_index:
            continue

        stock_code = str(item[0]).strip().upper()
        if not stock_code:
            continue

        rows.append(
            build_official_margin_row(
                symbol=f"{stock_code}.TW",
                record_date=normalized_date,
                current_balance=item[current_balance_index],
                previous_balance=item[previous_balance_index],
            )
        )

    if not rows:
        raise ValueError("證交所該日無可用融資餘額")

    return pd.DataFrame(rows, columns=MARGIN_REQUIRED_COLUMNS), normalized_date


def parse_tpex_margin_payload(payload: dict[str, Any]) -> tuple[pd.DataFrame, str]:
    """解析櫃買中心融資融券餘額資料。"""
    status_text = str(payload.get("stat", "")).strip().lower()
    if status_text != "ok":
        raise ValueError(str(payload.get("stat", "")).strip() or "櫃買中心端點未回傳有效資料")

    record_date = pd.to_datetime(payload.get("date"), format="%Y%m%d", errors="coerce")
    if pd.isna(record_date):
        raise ValueError("櫃買中心資料缺少有效日期")

    tables = payload.get("tables") or []
    if not tables:
        raise ValueError("櫃買中心該日無融資融券明細")

    detail_table = tables[0]
    if not detail_table.get("data"):
        raise ValueError("櫃買中心該日無融資融券明細")

    fields = detail_table.get("fields") or []
    previous_balance_index = fields.index("前資餘額(張)")
    current_balance_index = fields.index("資餘額")
    normalized_date = record_date.strftime("%Y-%m-%d")

    rows: list[dict[str, Any]] = []
    for item in detail_table.get("data", []):
        if len(item) <= current_balance_index:
            continue

        stock_code = str(item[0]).strip().upper()
        if not stock_code:
            continue

        rows.append(
            build_official_margin_row(
                symbol=f"{stock_code}.TWO",
                record_date=normalized_date,
                current_balance=item[current_balance_index],
                previous_balance=item[previous_balance_index],
            )
        )

    if not rows:
        raise ValueError("櫃買中心該日無可用融資餘額")

    return pd.DataFrame(rows, columns=MARGIN_REQUIRED_COLUMNS), normalized_date


def fetch_market_margin_data(
    market_label: str,
    reference_date: date,
    url_builder: Any,
    payload_parser: Any,
) -> dict[str, Any]:
    """往前回退到最近可用交易日，抓取單一市場的官方融資資料。"""
    recent_errors: list[str] = []

    for offset in range(OFFICIAL_MARGIN_LOOKBACK_DAYS + 1):
        target_date = reference_date - timedelta(days=offset)
        try:
            payload = fetch_json_from_url(url_builder(target_date))
            margin_df, used_date = payload_parser(payload)
            if not margin_df.empty:
                return {"data": margin_df, "used_date": used_date, "warning": ""}
        except Exception as exc:
            recent_errors.append(f"{target_date.isoformat()}：{exc}")

    summarized_errors = "；".join(recent_errors[-3:]) if recent_errors else "無額外錯誤訊息"
    return {
        "data": make_empty_margin_df(),
        "used_date": None,
        "warning": f"{market_label} 最近 {OFFICIAL_MARGIN_LOOKBACK_DAYS + 1} 天查無可用官方融資資料：{summarized_errors}",
    }


def download_official_margin_data_impl(reference_date_text: str | None = None) -> dict[str, Any]:
    """自動抓取證交所與櫃買中心最近可用的官方融資資料。"""
    reference_timestamp = pd.to_datetime(reference_date_text or date.today().isoformat(), errors="coerce")
    reference_day = reference_timestamp.date() if pd.notna(reference_timestamp) else date.today()

    market_configs = [
        ("上市", build_twse_margin_url, parse_twse_margin_payload),
        ("上櫃", build_tpex_margin_url, parse_tpex_margin_payload),
    ]

    frames: list[pd.DataFrame] = []
    source_parts: list[str] = []
    warnings: list[str] = []

    for market_label, url_builder, payload_parser in market_configs:
        market_result = fetch_market_margin_data(market_label, reference_day, url_builder, payload_parser)
        market_df = market_result["data"]
        if not market_df.empty:
            frames.append(market_df)
            if market_result["used_date"]:
                source_parts.append(f"{market_label} {market_result['used_date']}")
        elif market_result["warning"]:
            warnings.append(market_result["warning"])

    combined_df = pd.concat(frames, ignore_index=True) if frames else make_empty_margin_df()
    source_summary = "已自動抓取官方融資資料：" + "、".join(source_parts) if source_parts else ""

    fetch_error = ""
    if combined_df.empty:
        fetch_error = "目前無法自動取得官方融資資料。"
        if warnings:
            fetch_error = fetch_error + " " + "；".join(warnings)

    return {
        "data": combined_df,
        "source_summary": source_summary,
        "warnings": warnings,
        "fetch_error": fetch_error,
    }


@st.cache_data(ttl=OFFICIAL_MARGIN_CACHE_TTL_SECONDS, show_spinner=False)
def download_official_margin_data_cached(reference_date_text: str | None = None) -> dict[str, Any]:
    """快取官方融資資料，減少重複請求。"""
    return download_official_margin_data_impl(reference_date_text)


def should_replace_official_margin_result(cached_result: dict[str, Any], live_result: dict[str, Any]) -> bool:
    """比較快取與即時抓取結果，判斷是否應以即時結果覆蓋。"""
    cached_rows = len(cached_result.get("data", make_empty_margin_df()))
    live_rows = len(live_result.get("data", make_empty_margin_df()))
    cached_warning_count = len(cached_result.get("warnings", []))
    live_warning_count = len(live_result.get("warnings", []))

    if live_rows > cached_rows:
        return True
    if live_warning_count < cached_warning_count:
        return True
    if bool(cached_result.get("fetch_error")) and not live_result.get("fetch_error"):
        return True
    if not cached_result.get("source_summary") and live_result.get("source_summary"):
        return True
    return False


def download_official_margin_data(reference_date_text: str | None = None) -> dict[str, Any]:
    """取得官方融資資料；若快取只有部分市場成功，會再做一次即時重試。"""
    cached_result = download_official_margin_data_cached(reference_date_text)
    if not cached_result.get("warnings"):
        return cached_result

    live_result = download_official_margin_data_impl(reference_date_text)
    if should_replace_official_margin_result(cached_result, live_result):
        download_official_margin_data_cached.clear()
        return live_result
    return cached_result


def is_stop_broken(current_price: Any, final_trailing_stop: Any) -> bool:
    """判斷目前股價是否已跌破最終 ATR 移動停利價。"""
    current = to_float(current_price)
    stop = to_float(final_trailing_stop)
    return pd.notna(current) and pd.notna(stop) and current <= stop


def is_elevated_margin_risk(risk_text: Any) -> bool:
    """判斷是否屬於需要額外注意的融資風險。"""
    text = str(risk_text or "")
    keywords = ["風險升高", "惡化", "過熱", "提高移動停利紀律"]
    return any(keyword in text for keyword in keywords)


def determine_margin_risk_text(
    margin_change_rate: Any,
    foreign_buy_sell: Any,
    investment_trust_buy_sell: Any,
    price_change_rate: Any,
) -> str:
    """依融資增減、法人籌碼與股價漲跌幅回傳融資風險文字。"""
    margin_change_rate_value = to_float(margin_change_rate)
    foreign_buy_sell_value = to_float(foreign_buy_sell)
    investment_trust_buy_sell_value = to_float(investment_trust_buy_sell)
    price_change_rate_value = to_float(price_change_rate)

    risk_text = "融資變化中性，持續追蹤"
    if (
        pd.notna(foreign_buy_sell_value)
        and pd.notna(investment_trust_buy_sell_value)
        and pd.notna(margin_change_rate_value)
        and foreign_buy_sell_value < 0
        and investment_trust_buy_sell_value < 0
        and margin_change_rate_value > 0
    ):
        risk_text = "法人賣超但融資增加，散戶接籌碼風險升高"
    elif (
        pd.notna(margin_change_rate_value)
        and pd.notna(price_change_rate_value)
        and margin_change_rate_value > 5
        and price_change_rate_value < 0
    ):
        risk_text = "融資增加但股價下跌，籌碼惡化"
    elif (
        pd.notna(margin_change_rate_value)
        and pd.notna(price_change_rate_value)
        and margin_change_rate_value > 10
        and price_change_rate_value <= 3
    ):
        risk_text = "融資大增但股價漲幅有限，短線過熱"
    elif (
        pd.notna(margin_change_rate_value)
        and pd.notna(price_change_rate_value)
        and margin_change_rate_value > 10
        and price_change_rate_value > 5
    ):
        risk_text = "融資大增且股價強漲，續抱但需提高移動停利紀律"
    elif pd.notna(margin_change_rate_value) and margin_change_rate_value <= 0:
        risk_text = "融資未增加，籌碼壓力較低"

    return risk_text


def determine_risk_and_action(current_price: Any, final_trailing_stop: Any, cost: Any) -> tuple[str, str]:
    """依股價、成本與停利價關係回傳風險狀態與操作建議。"""
    current = to_float(current_price)
    stop = to_float(final_trailing_stop)
    cost_value = to_float(cost)

    if pd.isna(current) or pd.isna(stop) or pd.isna(cost_value):
        return RISK_DATA_INSUFFICIENT, "資料不足，暫不操作"

    if current <= stop and current > cost_value:
        return "跌破ATR移動停利，建議分批停利", "建議分批停利或至少減碼"

    if current <= stop and current <= cost_value:
        return "低於成本且跌破ATR防守，建議減碼或停損", "建議減碼或停損，避免虧損擴大"

    if current > stop and current > cost_value:
        return "獲利中，尚未跌破移動停利，續抱觀察", "續抱，停利線只上移不下移"

    return "低於成本但尚未跌破ATR防守，觀察反彈", "暫不加碼，觀察是否站回成本"


# 1. 讀取使用者上傳的 CSV。
def load_uploaded_csv(uploaded_file: Any) -> pd.DataFrame:
    """讀取使用者上傳的 CSV，優先支援 UTF-8 與常見繁中編碼。"""
    if uploaded_file is None:
        return pd.DataFrame()

    file_bytes = uploaded_file.getvalue()
    if not file_bytes:
        raise ValueError("上傳的 CSV 內容是空的。")

    for encoding in ("utf-8-sig", "utf-8", "cp950", "big5"):
        try:
            return pd.read_csv(BytesIO(file_bytes), encoding=encoding)
        except UnicodeDecodeError:
            continue
        except pd.errors.EmptyDataError as exc:
            raise ValueError("上傳的 CSV 內容是空的。") from exc

    raise ValueError("無法解析 CSV，請確認檔案為 UTF-8 或 Big5 編碼。")


# 2. 載入專案內建的 portfolio 樣板檔。
def load_sample_portfolio() -> pd.DataFrame:
    """載入專案提供的 portfolio_sample.csv，方便使用者下載修改。"""
    return pd.read_csv(PORTFOLIO_SAMPLE_PATH)


def clean_ocr_text(text: Any) -> str:
    """清理 OCR 文字，將全形數字與標點統一成可解析格式。"""
    normalized = unicodedata.normalize("NFKC", str(text or ""))
    normalized = normalized.replace("|", " ").replace("_", " ").replace("\t", " ")
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip()


@st.cache_resource(show_spinner=False)
def get_ocr_engine() -> Any:
    """延遲載入 OCR 模型，避免在未使用截圖功能時增加啟動成本。"""
    from rapidocr_onnxruntime import RapidOCR

    return RapidOCR()


def extract_text_lines_from_ocr_result(ocr_result: Any) -> list[str]:
    """將 OCR 偵測框依垂直位置合併成較容易解析的逐行文字。"""
    if not ocr_result:
        return []

    detections: list[dict[str, Any]] = []
    for item in ocr_result:
        if len(item) < 3:
            continue
        box, text, score = item[0], item[1], item[2]
        if not text or score < 0.35:
            continue
        x_positions = [point[0] for point in box]
        y_positions = [point[1] for point in box]
        detections.append(
            {
                "text": clean_ocr_text(text),
                "x": min(x_positions),
                "y": min(y_positions),
                "y_center": sum(y_positions) / len(y_positions),
                "height": max(y_positions) - min(y_positions),
            }
        )

    if not detections:
        return []

    detections.sort(key=lambda item: (item["y"], item["x"]))
    median_height = float(pd.Series([item["height"] for item in detections]).median())
    line_gap_threshold = max(12.0, median_height * 0.6)

    grouped_lines: list[list[dict[str, Any]]] = []
    current_group: list[dict[str, Any]] = []
    current_y_center = 0.0
    for detection in detections:
        if not current_group:
            current_group = [detection]
            current_y_center = detection["y_center"]
            continue

        if abs(detection["y_center"] - current_y_center) <= line_gap_threshold:
            current_group.append(detection)
            current_y_center = sum(item["y_center"] for item in current_group) / len(current_group)
        else:
            grouped_lines.append(sorted(current_group, key=lambda item: item["x"]))
            current_group = [detection]
            current_y_center = detection["y_center"]

    if current_group:
        grouped_lines.append(sorted(current_group, key=lambda item: item["x"]))

    return [clean_ocr_text(" ".join(item["text"] for item in group if item["text"])) for group in grouped_lines]


def first_non_empty(series: pd.Series, default_value: Any = "") -> Any:
    """回傳序列中第一個非空值；若沒有則使用預設值。"""
    for value in series:
        if pd.notna(value) and str(value).strip() not in {"", "nan", "None"}:
            return value
    return default_value


def parse_portfolio_candidate_line(
    line: str,
    default_category: str,
    share_unit_multiplier: int,
    image_name: str,
) -> dict[str, Any] | None:
    """從單行 OCR 文字中盡量推估股票代號、名稱、股數與成本價。"""
    cleaned_line = clean_ocr_text(line)
    if not cleaned_line:
        return None

    if any(keyword in cleaned_line for keyword in OCR_HEADER_KEYWORDS):
        return None

    name_before_code_match = re.search(r"(?P<name>.+?)[（(](?P<code>\d{4,6})[）)]", cleaned_line)
    leading_name = ""

    if name_before_code_match:
        symbol = normalize_symbol(name_before_code_match.group("code"))
        leading_name = clean_ocr_text(name_before_code_match.group("name")).strip(" :-|")
        trailing_text = cleaned_line[name_before_code_match.end() :].strip(" :-|")
    else:
        code_match = re.search(r"(?<!\d)(\d{4,6})(?!\d)", cleaned_line)
        if not code_match:
            return None

        symbol = normalize_symbol(code_match.group(1))
        leading_name = clean_ocr_text(cleaned_line[: code_match.start()])
        leading_name = re.sub(r"[（(]+$", "", leading_name).strip(" :-|")
        trailing_text = cleaned_line[code_match.end() :].lstrip(")）").strip(" :-|")

    tokenized_parts = [part for part in re.split(r"\s+", trailing_text) if part]
    name_tokens: list[str] = []
    numeric_tokens: list[str] = []

    for part in tokenized_parts:
        normalized_part = part.replace(",", "").rstrip("%")
        is_numeric_token = bool(re.fullmatch(r"-?\d+(?:\.\d+)?", normalized_part))

        if part not in OCR_NON_NAME_TOKENS and not is_numeric_token and not leading_name:
            name_tokens.append(part)

        if is_numeric_token and not normalized_part.startswith("-"):
            numeric_tokens.append(normalized_part)

        if part in OCR_NON_NAME_TOKENS:
            continue

        if is_numeric_token and not leading_name:
            break

    name = leading_name or " ".join(name_tokens).strip(" :-|")

    integer_values: list[float] = []
    decimal_values: list[float] = []
    for token in numeric_tokens:
        if len(token) >= 8 and "." not in token:
            continue
        numeric_value = float(token)
        if "." in token:
            decimal_values.append(numeric_value)
        else:
            integer_values.append(numeric_value)

    shares_value: Any = pd.NA
    cost_value: Any = pd.NA

    if integer_values:
        share_candidate = integer_values[0]
        if not decimal_values and len(integer_values) >= 2 and integer_values[0] <= 5000 < integer_values[1]:
            share_candidate = integer_values[1]
            if not decimal_values:
                cost_value = float(integer_values[0])
        shares_value = share_candidate * share_unit_multiplier

    if decimal_values:
        cost_value = float(decimal_values[0])
    elif len(integer_values) >= 2 and pd.isna(cost_value):
        cost_candidates = [value for value in integer_values if 0 < value <= 5000]
        if cost_candidates:
            cost_value = float(min(cost_candidates))

    if not name:
        name = symbol.replace(".TW", "")

    return {
        "symbol": symbol,
        "name": name,
        "shares": shares_value,
        "cost": cost_value,
        "category": default_category,
        "image_name": image_name,
        "ocr_text": cleaned_line,
    }


def consolidate_ocr_portfolio_df(portfolio_df: pd.DataFrame, default_category: str) -> pd.DataFrame:
    """合併多張截圖中的重複股票，並保留原始 OCR 文字供人工校對。"""
    if portfolio_df.empty:
        return make_empty_ocr_portfolio_df()

    working_df = portfolio_df.copy()
    working_df["symbol"] = working_df["symbol"].map(normalize_symbol)
    working_df["shares"] = pd.to_numeric(working_df["shares"], errors="coerce")
    working_df["cost"] = pd.to_numeric(working_df["cost"], errors="coerce")

    consolidated_rows: list[dict[str, Any]] = []
    for symbol, group in working_df.groupby("symbol", sort=True, dropna=False):
        if not symbol:
            continue
        consolidated_rows.append(
            {
                "symbol": symbol,
                "name": first_non_empty(group["name"], symbol.replace(".TW", "")),
                "shares": group["shares"].sum(min_count=1),
                "cost": first_non_empty(group["cost"], pd.NA),
                "category": first_non_empty(group["category"], default_category),
                "image_name": ", ".join(sorted({str(value) for value in group["image_name"] if str(value).strip()})),
                "ocr_text": " | ".join(str(value) for value in group["ocr_text"] if str(value).strip()),
            }
        )

    return pd.DataFrame(consolidated_rows, columns=OCR_PORTFOLIO_COLUMNS)


def extract_portfolio_candidates_from_images(
    uploaded_files: list[Any],
    default_category: str,
    share_unit_multiplier: int,
) -> pd.DataFrame:
    """對持股截圖執行 OCR，產出可人工確認的持股草稿。"""
    ocr_engine = get_ocr_engine()
    parsed_rows: list[dict[str, Any]] = []
    raw_lines: list[str] = []

    for uploaded_file in uploaded_files:
        image = Image.open(BytesIO(uploaded_file.getvalue())).convert("RGB")
        processed_image = ImageOps.autocontrast(ImageOps.grayscale(image))
        ocr_result, _ = ocr_engine(np.array(processed_image))
        line_texts = extract_text_lines_from_ocr_result(ocr_result)
        raw_lines.extend([f"[{uploaded_file.name}] {line}" for line in line_texts])

        for line in line_texts:
            parsed_row = parse_portfolio_candidate_line(
                line,
                default_category=default_category,
                share_unit_multiplier=share_unit_multiplier,
                image_name=uploaded_file.name,
            )
            if parsed_row:
                parsed_rows.append(parsed_row)

    portfolio_df = pd.DataFrame(parsed_rows, columns=OCR_PORTFOLIO_COLUMNS)
    portfolio_df = consolidate_ocr_portfolio_df(portfolio_df, default_category)
    portfolio_df.attrs["raw_lines"] = raw_lines
    return portfolio_df


def finalize_portfolio_input(portfolio_df: pd.DataFrame, default_category: str = "normal") -> pd.DataFrame:
    """將 CSV 或 OCR 草稿整理成可供報表計算的標準持股資料表。"""
    if portfolio_df.empty:
        return pd.DataFrame(columns=PORTFOLIO_REQUIRED_COLUMNS)

    working_df = portfolio_df.copy()
    for column in PORTFOLIO_REQUIRED_COLUMNS:
        if column not in working_df.columns:
            working_df[column] = pd.NA

    working_df["symbol"] = working_df["symbol"].fillna("").astype(str).str.strip()
    working_df["name"] = working_df["name"].fillna("").astype(str).str.strip()
    working_df.loc[working_df["name"] == "", "name"] = working_df.loc[working_df["name"] == "", "symbol"]
    working_df["shares"] = pd.to_numeric(working_df["shares"], errors="coerce")
    working_df["cost"] = pd.to_numeric(working_df["cost"], errors="coerce")
    working_df["category"] = (
        working_df["category"].fillna(default_category).astype(str).str.strip().replace("", default_category)
    )
    working_df = working_df[
        ~(
            working_df["symbol"].eq("")
            & working_df["name"].eq("")
            & working_df["shares"].isna()
            & working_df["cost"].isna()
        )
    ]
    return working_df


# 3. 載入融資資料；若未上傳則改抓證交所與櫃買中心官方資料。
def load_margin_data(uploaded_file: Any, reference_date: date | None = None) -> pd.DataFrame:
    """載入使用者上傳的融資資料，或自動抓取官方融資資料。"""
    if uploaded_file is None:
        official_result = download_official_margin_data((reference_date or date.today()).isoformat())
        margin_df = official_result["data"].copy()
        margin_df.attrs["source"] = "official"
        margin_df.attrs["source_summary"] = official_result.get("source_summary", "")
        margin_df.attrs["fetch_warnings"] = official_result.get("warnings", [])
        margin_df.attrs["fetch_error"] = official_result.get("fetch_error", "")
        return margin_df

    try:
        margin_df = load_uploaded_csv(uploaded_file)
    except ValueError as exc:
        margin_df = make_empty_margin_df()
        margin_df.attrs["source"] = "upload"
        margin_df.attrs["validation_error"] = str(exc)
        return margin_df

    missing_columns = validate_required_columns(margin_df, MARGIN_REQUIRED_COLUMNS)
    if missing_columns:
        empty_df = make_empty_margin_df()
        empty_df.attrs["source"] = "upload"
        empty_df.attrs["missing_columns"] = missing_columns
        return empty_df

    margin_df = margin_df.copy()
    margin_df["symbol"] = margin_df["symbol"].map(normalize_symbol)
    margin_df["date"] = pd.to_datetime(margin_df["date"], errors="coerce")

    for column in [
        "margin_balance",
        "margin_change",
        "margin_change_rate",
        "foreign_buy_sell",
        "investment_trust_buy_sell",
        "price_change_rate",
    ]:
        margin_df[column] = pd.to_numeric(margin_df[column], errors="coerce")

    margin_df = margin_df.dropna(subset=["symbol"])
    margin_df["date"] = margin_df["date"].dt.strftime("%Y-%m-%d")
    margin_df.attrs["source"] = "upload"
    return margin_df


# 4. 載入停利歷史檔，若未提供則從空白歷史開始。
def load_stop_history(uploaded_file: Any) -> pd.DataFrame:
    """載入 stop_history.csv，並把歷史停利價整理成可比較的數值欄位。"""
    if uploaded_file is None:
        return make_empty_stop_history_df()

    try:
        stop_history_df = load_uploaded_csv(uploaded_file)
    except ValueError as exc:
        stop_history_df = make_empty_stop_history_df()
        stop_history_df.attrs["validation_error"] = str(exc)
        return stop_history_df

    missing_columns = validate_required_columns(stop_history_df, STOP_HISTORY_REQUIRED_COLUMNS)
    if missing_columns:
        empty_df = make_empty_stop_history_df()
        empty_df.attrs["missing_columns"] = missing_columns
        return empty_df

    stop_history_df = stop_history_df.copy()
    stop_history_df["symbol"] = stop_history_df["symbol"].map(normalize_symbol)
    stop_history_df["last_trailing_stop"] = pd.to_numeric(
        stop_history_df["last_trailing_stop"], errors="coerce"
    )
    stop_history_df["last_update_date"] = pd.to_datetime(
        stop_history_df["last_update_date"], errors="coerce"
    ).dt.strftime("%Y-%m-%d")
    stop_history_df = stop_history_df.dropna(subset=["symbol"])
    return stop_history_df


# 5. 下載單一股票日線資料，並以 Streamlit 快取減少重複抓取。
def download_stock_data_impl(symbol: str) -> dict[str, Any]:
    """從 yfinance 下載單一股票最近至少 120 天以上的日線 OHLCV 資料。"""
    candidate_symbols = build_symbol_lookup_candidates(symbol)
    end_date = pd.Timestamp.today().normalize() + pd.Timedelta(days=1)
    start_date = end_date - pd.Timedelta(days=365)
    required_columns = ["Open", "High", "Low", "Close", "Volume"]
    failure_messages: list[str] = []

    for candidate_symbol in candidate_symbols:
        try:
            stock_df = yf.download(
                candidate_symbol,
                start=start_date,
                end=end_date,
                interval="1d",
                progress=False,
                auto_adjust=False,
                threads=False,
            )
        except Exception as exc:
            failure_messages.append(f"{candidate_symbol}: {exc}")
            continue

        if stock_df is None or stock_df.empty:
            failure_messages.append(f"{candidate_symbol}: 查無有效日線資料。")
            continue

        if isinstance(stock_df.columns, pd.MultiIndex):
            stock_df.columns = stock_df.columns.get_level_values(0)

        missing_columns = validate_required_columns(stock_df, required_columns)
        if missing_columns:
            failure_messages.append(f"{candidate_symbol}: 缺少欄位：{', '.join(missing_columns)}")
            continue

        stock_df = stock_df[required_columns].copy()
        stock_df = stock_df.dropna(subset=["High", "Low", "Close"])
        stock_df.index = pd.to_datetime(stock_df.index)
        return {
            "success": True,
            "message": "",
            "data": stock_df,
            "symbol": candidate_symbol,
        }

    return {
        "success": False,
        "message": "；".join(failure_messages) if failure_messages else "查無有效日線資料。",
        "data": pd.DataFrame(),
        "symbol": normalize_symbol(symbol),
    }


@st.cache_data(ttl=3600, show_spinner=False)
def download_stock_data_cached(symbol: str) -> dict[str, Any]:
    """快取成功的股價下載結果，降低重複請求頻率。"""
    return download_stock_data_impl(symbol)


def download_stock_data(symbol: str) -> dict[str, Any]:
    """取得股價資料；若快取命中的是失敗結果，會即時重試一次。"""
    cached_result = download_stock_data_cached(symbol)
    if cached_result.get("success"):
        return cached_result

    live_result = download_stock_data_impl(symbol)
    if live_result.get("success"):
        download_stock_data_cached.clear()
        return live_result
    return live_result


# 6. 計算標準 ATR 所需的 TR 與 ATR 欄位。
def calculate_atr(stock_df: pd.DataFrame, atr_period: int) -> pd.DataFrame:
    """依標準公式計算每日 TR 與 ATR，回傳附加欄位後的資料表。"""
    atr_df = stock_df.copy()
    if atr_df.empty:
        return atr_df

    atr_df["PreviousClose"] = atr_df["Close"].shift(1)
    atr_df["TR"] = pd.concat(
        [
            atr_df["High"] - atr_df["Low"],
            (atr_df["High"] - atr_df["PreviousClose"]).abs(),
            (atr_df["Low"] - atr_df["PreviousClose"]).abs(),
        ],
        axis=1,
    ).max(axis=1)
    atr_df.loc[atr_df["PreviousClose"].isna(), "TR"] = atr_df["High"] - atr_df["Low"]
    atr_df["ATR"] = atr_df["TR"].rolling(window=atr_period, min_periods=atr_period).mean()
    return atr_df


def calculate_daily_price_change_rate(stock_df: pd.DataFrame) -> float:
    """使用最近兩個交易日收盤價計算單日漲跌幅。"""
    if stock_df.empty or "Close" not in stock_df.columns or len(stock_df) < 2:
        return float("nan")

    latest_close = to_float(stock_df["Close"].iloc[-1])
    previous_close = to_float(stock_df["Close"].iloc[-2])
    if pd.isna(latest_close) or pd.isna(previous_close) or previous_close == 0:
        return float("nan")

    return ((latest_close - previous_close) / previous_close) * 100


# 7. 依照持股類型回傳 ATR 倍數，並允許使用者用側邊欄手動覆寫。
def get_atr_multiplier(category: Any, multiplier_settings: dict[str, float]) -> float:
    """根據持股類型決定 ATR 倍數；找不到時回傳預設值 2.0。"""
    category_text = str(category).strip()
    if category_text in multiplier_settings:
        return float(multiplier_settings[category_text])
    return 2.0


# 8. 計算 recent high、最新 TR、ATR 與原始移動停利價。
def calculate_trailing_stop(
    stock_df: pd.DataFrame, atr_multiplier: float, high_period: int
) -> dict[str, Any]:
    """使用最近高點與 ATR 倍數計算原始 ATR 移動停利價。"""
    if stock_df.empty or "ATR" not in stock_df.columns:
        return {"success": False, "message": RISK_DATA_INSUFFICIENT}

    latest_atr = stock_df["ATR"].iloc[-1]
    latest_tr = stock_df["TR"].iloc[-1] if "TR" in stock_df.columns else float("nan")
    if len(stock_df) < high_period or pd.isna(latest_atr):
        return {
            "success": False,
            "message": RISK_DATA_INSUFFICIENT,
            "latest_tr": latest_tr,
            "atr": latest_atr,
        }

    recent_window = stock_df.tail(high_period)
    recent_high = recent_window["High"].max()
    raw_trailing_stop = recent_high - (latest_atr * atr_multiplier)
    return {
        "success": True,
        "message": "",
        "recent_high": recent_high,
        "latest_tr": latest_tr,
        "atr": latest_atr,
        "raw_trailing_stop": raw_trailing_stop,
    }


# 9. 根據舊 stop history 決定本次最終停利價，確保停利線只能上移。
def update_final_trailing_stop(
    symbol: str, name: str, raw_trailing_stop: Any, stop_history_df: pd.DataFrame
) -> dict[str, Any]:
    """比較新舊停利價，若歷史停利價較高則維持歷史值。"""
    del name
    raw_stop = to_float(raw_trailing_stop)
    if pd.isna(raw_stop):
        return {"last_trailing_stop": float("nan"), "final_trailing_stop": float("nan")}

    if stop_history_df.empty:
        return {"last_trailing_stop": float("nan"), "final_trailing_stop": raw_stop}

    normalized_symbol = normalize_symbol(symbol)
    matched = stop_history_df[stop_history_df["symbol"] == normalized_symbol].copy()
    if matched.empty:
        symbol_code = get_symbol_code(normalized_symbol)
        matched = stop_history_df[stop_history_df["symbol"].map(get_symbol_code) == symbol_code].copy()
    if matched.empty:
        return {"last_trailing_stop": float("nan"), "final_trailing_stop": raw_stop}

    last_trailing_stop = pd.to_numeric(matched["last_trailing_stop"], errors="coerce").max()
    if pd.isna(last_trailing_stop):
        return {"last_trailing_stop": float("nan"), "final_trailing_stop": raw_stop}

    return {
        "last_trailing_stop": last_trailing_stop,
        "final_trailing_stop": max(raw_stop, float(last_trailing_stop)),
    }


# 10. 依照融資資料規則判斷單一股票的籌碼風險。
def calculate_margin_risk(symbol: str, margin_df: pd.DataFrame) -> dict[str, Any]:
    """整合融資與法人資料，回傳融資風險文字與相關欄位值。"""
    result = {
        "融資餘額": pd.NA,
        "融資增減": pd.NA,
        "融資增減率": pd.NA,
        "外資買賣超": pd.NA,
        "投信買賣超": pd.NA,
        "股價漲跌幅": pd.NA,
        "融資風險": "尚未提供融資資料",
        "margin_date": None,
    }

    if margin_df.empty:
        return result

    normalized_symbol = normalize_symbol(symbol)
    matched = margin_df[margin_df["symbol"] == normalized_symbol].copy()
    if matched.empty:
        symbol_code = get_symbol_code(normalized_symbol)
        matched = margin_df[margin_df["symbol"].map(get_symbol_code) == symbol_code].copy()
    if matched.empty:
        return result

    matched["date"] = pd.to_datetime(matched["date"], errors="coerce")
    matched = matched.sort_values("date")
    row = matched.iloc[-1]

    risk_text = determine_margin_risk_text(
        row.get("margin_change_rate"),
        row.get("foreign_buy_sell"),
        row.get("investment_trust_buy_sell"),
        row.get("price_change_rate"),
    )

    result.update(
        {
            "融資餘額": row.get("margin_balance", pd.NA),
            "融資增減": row.get("margin_change", pd.NA),
            "融資增減率": row.get("margin_change_rate", pd.NA),
            "外資買賣超": row.get("foreign_buy_sell", pd.NA),
            "投信買賣超": row.get("investment_trust_buy_sell", pd.NA),
            "股價漲跌幅": row.get("price_change_rate", pd.NA),
            "融資風險": risk_text,
            "margin_date": row.get("date"),
        }
    )
    return result


# 11. 整合 portfolio、股價、停利歷史與融資資料，產出最終中文報表。
def build_report(
    portfolio_df: pd.DataFrame,
    margin_df: pd.DataFrame,
    stop_history_df: pd.DataFrame,
    settings: dict[str, Any],
) -> pd.DataFrame:
    """建立完整 ATR 報表，單檔失敗不影響整體結果。"""
    report_rows: list[dict[str, Any]] = []
    processing_messages: list[str] = []

    working_portfolio = portfolio_df.copy()
    working_portfolio["symbol"] = working_portfolio["symbol"].map(normalize_symbol)

    for _, row in working_portfolio.iterrows():
        symbol = normalize_symbol(row.get("symbol", ""))
        name = str(row.get("name", "")).strip()
        shares = to_float(row.get("shares"))
        cost = to_float(row.get("cost"))
        category = row.get("category", "normal")
        atr_multiplier = get_atr_multiplier(category, settings["multiplier_settings"])
        cost_value = shares * cost if pd.notna(shares) and pd.notna(cost) else float("nan")

        report_row = {column: pd.NA for column in REPORT_COLUMNS}
        report_row.update(
            {
                "股票代號": symbol,
                "股票名稱": name,
                "股數": shares,
                "成本價": cost,
                "投入成本": cost_value,
                "ATR倍數": atr_multiplier,
                "風險狀態": RISK_DATA_INSUFFICIENT,
                "操作建議": "資料不足，暫不操作",
                "融資風險": "尚未提供融資資料",
                "資料更新日期": date.today().isoformat(),
            }
        )

        margin_info = calculate_margin_risk(symbol, margin_df)
        for field in ["融資餘額", "融資增減", "融資增減率", "外資買賣超", "投信買賣超", "股價漲跌幅", "融資風險"]:
            report_row[field] = margin_info[field]

        margin_date = margin_info.get("margin_date")
        if pd.notna(margin_date):
            report_row["資料更新日期"] = pd.to_datetime(margin_date).strftime("%Y-%m-%d")

        if not symbol or pd.isna(shares) or pd.isna(cost):
            report_row["風險狀態"] = "持股資料格式錯誤，暫不判斷"
            report_row["操作建議"] = "請修正 portfolio.csv 後重新上傳"
            report_rows.append(report_row)
            continue

        download_result = download_stock_data(symbol)
        if not download_result["success"]:
            report_row["風險狀態"] = RISK_DOWNLOAD_FAILURE
            report_row["操作建議"] = "資料不足，暫不操作"
            processing_messages.append(f"{symbol} 股價下載失敗：{download_result['message']}")
            report_rows.append(report_row)
            continue

        symbol = normalize_symbol(download_result.get("symbol", symbol))
        report_row["股票代號"] = symbol
        stock_df = download_result["data"]
        if stock_df.empty:
            report_row["風險狀態"] = RISK_DOWNLOAD_FAILURE
            report_row["操作建議"] = "資料不足，暫不操作"
            processing_messages.append(f"{symbol} 股價下載失敗：回傳空白資料。")
            report_rows.append(report_row)
            continue

        latest_close = to_float(stock_df["Close"].iloc[-1])
        market_value = latest_close * shares if pd.notna(latest_close) and pd.notna(shares) else float("nan")
        unrealized_profit = market_value - cost_value if pd.notna(market_value) and pd.notna(cost_value) else float("nan")
        unrealized_profit_rate = (
            (unrealized_profit / cost_value) * 100
            if pd.notna(unrealized_profit) and pd.notna(cost_value) and cost_value != 0
            else float("nan")
        )

        report_row.update(
            {
                "現價": latest_close,
                "市值": market_value,
                "未實現損益": unrealized_profit,
                "未實現損益率": unrealized_profit_rate,
                "資料更新日期": stock_df.index[-1].strftime("%Y-%m-%d"),
            }
        )

        if pd.isna(to_float(report_row["股價漲跌幅"])):
            computed_price_change_rate = calculate_daily_price_change_rate(stock_df)
            if pd.notna(computed_price_change_rate):
                report_row["股價漲跌幅"] = computed_price_change_rate

        if pd.notna(to_float(report_row["融資餘額"])):
            report_row["融資風險"] = determine_margin_risk_text(
                report_row["融資增減率"],
                report_row["外資買賣超"],
                report_row["投信買賣超"],
                report_row["股價漲跌幅"],
            )

        atr_df = calculate_atr(stock_df, settings["atr_period"])
        trailing_stop_result = calculate_trailing_stop(atr_df, atr_multiplier, settings["high_period"])

        if not trailing_stop_result["success"]:
            report_row["最新TR"] = trailing_stop_result.get("latest_tr", pd.NA)
            report_row["ATR"] = trailing_stop_result.get("atr", pd.NA)
            report_row["風險狀態"] = RISK_DATA_INSUFFICIENT
            report_row["操作建議"] = "資料不足，暫不操作"
            processing_messages.append(f"{symbol} 資料不足：ATR 週期或最近高點期間不足。")
            report_rows.append(report_row)
            continue

        raw_trailing_stop = trailing_stop_result["raw_trailing_stop"]
        stop_update = update_final_trailing_stop(symbol, name, raw_trailing_stop, stop_history_df)
        final_trailing_stop = stop_update["final_trailing_stop"]
        distance_to_stop = (
            latest_close - final_trailing_stop
            if pd.notna(latest_close) and pd.notna(final_trailing_stop)
            else float("nan")
        )
        distance_to_stop_rate = (
            (distance_to_stop / latest_close) * 100
            if pd.notna(distance_to_stop) and pd.notna(latest_close) and latest_close != 0
            else float("nan")
        )
        risk_status, action_suggestion = determine_risk_and_action(latest_close, final_trailing_stop, cost)

        report_row.update(
            {
                "最近高點": trailing_stop_result["recent_high"],
                "最新TR": trailing_stop_result["latest_tr"],
                "ATR": trailing_stop_result["atr"],
                "原始ATR移動停利價": raw_trailing_stop,
                "最終ATR移動停利價": final_trailing_stop,
                "距離停利價差": distance_to_stop,
                "距離停利百分比": distance_to_stop_rate,
                "風險狀態": risk_status,
                "操作建議": action_suggestion,
            }
        )
        report_rows.append(report_row)

    report_df = pd.DataFrame(report_rows, columns=REPORT_COLUMNS)
    report_df.attrs["processing_messages"] = processing_messages
    return report_df


# 12. 依本次報表結果產出新的 stop_history.csv，供使用者下次再上傳。
def build_updated_stop_history(
    report_df: pd.DataFrame, old_stop_history_df: pd.DataFrame
) -> pd.DataFrame:
    """以本次報表中的最終 ATR 停利價更新 stop history。"""
    rows: list[dict[str, Any]] = []

    for _, row in report_df.iterrows():
        symbol = row["股票代號"]
        name = row["股票名稱"]
        new_stop = to_float(row["最終ATR移動停利價"])
        existing = old_stop_history_df[old_stop_history_df["symbol"] == symbol].copy()

        old_stop = pd.to_numeric(existing.get("last_trailing_stop"), errors="coerce").max() if not existing.empty else float("nan")
        old_update_date = None
        if not existing.empty:
            existing = existing.sort_values("last_update_date")
            old_update_date = existing.iloc[-1].get("last_update_date")

        final_stop = new_stop
        update_date = row["資料更新日期"]
        if pd.isna(final_stop) and pd.notna(old_stop):
            final_stop = float(old_stop)
            update_date = old_update_date

        if pd.isna(final_stop):
            continue

        rows.append(
            {
                "symbol": symbol,
                "name": name,
                "last_trailing_stop": final_stop,
                "last_update_date": update_date,
            }
        )

    updated_stop_history_df = pd.DataFrame(rows, columns=STOP_HISTORY_REQUIRED_COLUMNS)
    if updated_stop_history_df.empty:
        return updated_stop_history_df

    updated_stop_history_df = updated_stop_history_df.sort_values("symbol").drop_duplicates(
        subset=["symbol"], keep="last"
    )
    return updated_stop_history_df.reset_index(drop=True)


# 13. 產生 Excel 下載內容並套用格式樣式。
def create_excel_bytes(report_df: pd.DataFrame) -> bytes:
    """使用 openpyxl 產生符合條件格式要求的 Excel 檔案。"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ATR報表"

    worksheet.append(REPORT_COLUMNS)
    for row in report_df[REPORT_COLUMNS].itertuples(index=False, name=None):
        worksheet.append([safe_excel_value(value) for value in row])

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    stop_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    margin_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    red_font = Font(color="C00000")
    green_font = Font(color="008000")

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    column_index = {column: index + 1 for index, column in enumerate(REPORT_COLUMNS)}

    for row_index in range(2, worksheet.max_row + 1):
        current_price = worksheet.cell(row=row_index, column=column_index["現價"]).value
        final_stop = worksheet.cell(row=row_index, column=column_index["最終ATR移動停利價"]).value
        if is_stop_broken(current_price, final_stop):
            for column_number in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_index, column=column_number).fill = stop_fill

        profit_cell = worksheet.cell(row=row_index, column=column_index["未實現損益"])
        profit_rate_cell = worksheet.cell(row=row_index, column=column_index["未實現損益率"])
        profit_value = to_float(profit_cell.value)
        if pd.notna(profit_value):
            if profit_value > 0:
                profit_cell.font = red_font
                profit_rate_cell.font = red_font
            elif profit_value < 0:
                profit_cell.font = green_font
                profit_rate_cell.font = green_font

        margin_risk_cell = worksheet.cell(row=row_index, column=column_index["融資風險"])
        if is_elevated_margin_risk(margin_risk_cell.value):
            margin_risk_cell.fill = margin_fill

        for column_name in DISPLAY_PERCENT_COLUMNS:
            cell = worksheet.cell(row=row_index, column=column_index[column_name])
            if cell.value is not None:
                cell.value = to_float(cell.value) / 100 if pd.notna(to_float(cell.value)) else None
                cell.number_format = "0.00%"

        for column_name in DISPLAY_NUMBER_COLUMNS:
            cell = worksheet.cell(row=row_index, column=column_index[column_name])
            if cell.value is not None:
                cell.number_format = "#,##0.00"

        shares_cell = worksheet.cell(row=row_index, column=column_index["股數"])
        if shares_cell.value is not None:
            shares_cell.number_format = "#,##0"

    for column_number, column_name in enumerate(REPORT_COLUMNS, start=1):
        max_length = len(str(column_name))
        for row in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row=row, column=column_number).value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[get_column_letter(column_number)].width = min(max_length + 2, 36)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


# 14. 產生報表 CSV 下載內容。
def create_csv_bytes(report_df: pd.DataFrame) -> bytes:
    """將報表輸出為 UTF-8 BOM CSV，方便 Excel 開啟中文欄位。"""
    return report_df[REPORT_COLUMNS].to_csv(index=False).encode("utf-8-sig")


# 15. 產生 updated stop history CSV 下載內容。
def create_stop_history_csv_bytes(stop_history_df: pd.DataFrame) -> bytes:
    """輸出更新後的 stop history CSV，讓使用者下次可再上傳。"""
    return stop_history_df[STOP_HISTORY_REQUIRED_COLUMNS].to_csv(index=False).encode("utf-8-sig")


def style_report_dataframe(report_df: pd.DataFrame) -> pd.io.formats.style.Styler:
    """把報表套用 Streamlit 表格格式，包含紅綠損益與停利提示。"""
    display_df = report_df.copy()

    number_formats = {column: "{:,.2f}" for column in DISPLAY_NUMBER_COLUMNS if column in display_df.columns}
    if "股數" in display_df.columns:
        number_formats["股數"] = "{:,.0f}"
    number_formats.update({column: "{:.2f}%" for column in DISPLAY_PERCENT_COLUMNS if column in display_df.columns})

    def apply_row_style(row: pd.Series) -> list[str]:
        if not is_stop_broken(row.get("現價"), row.get("最終ATR移動停利價")):
            return [""] * len(row)

        row_styles = ["border-top: 2px solid #F59E0B; border-bottom: 2px solid #F59E0B;"] * len(row)
        if row_styles:
            row_styles[0] += " border-left: 4px solid #F59E0B;"
            row_styles[-1] += " border-right: 4px solid #F59E0B;"
        return row_styles

    def profit_color(value: Any) -> str:
        numeric_value = to_float(value)
        if pd.isna(numeric_value):
            return ""
        if numeric_value > 0:
            return "color: #C00000; font-weight: 600"
        if numeric_value < 0:
            return "color: #008000; font-weight: 600"
        return ""

    def margin_risk_style(value: Any) -> str:
        return "background-color: #F4CCCC" if is_elevated_margin_risk(value) else ""

    styler = display_df.style.format(number_formats, na_rep="-").apply(apply_row_style, axis=1)

    profit_subset = [column for column in ["未實現損益", "未實現損益率"] if column in display_df.columns]
    if profit_subset:
        styler = styler.map(profit_color, subset=profit_subset)

    if "融資風險" in display_df.columns:
        styler = styler.map(margin_risk_style, subset=["融資風險"])

    return styler


# 16. 呈現主畫面的指標卡、篩選器、完整報表與注意清單。
def render_dashboard(report_df: pd.DataFrame) -> None:
    """把計算完成的報表渲染成可搜尋、可篩選的儀表板。"""
    if report_df.empty:
        st.info("目前沒有可顯示的報表資料。")
        return

    market_value_total = pd.to_numeric(report_df["市值"], errors="coerce").sum()
    cost_total = pd.to_numeric(report_df["投入成本"], errors="coerce").sum()
    unrealized_profit_total = pd.to_numeric(report_df["未實現損益"], errors="coerce").sum()
    unrealized_profit_rate_total = (
        (unrealized_profit_total / cost_total) * 100 if pd.notna(cost_total) and cost_total != 0 else 0.0
    )
    stop_break_count = int(
        report_df.apply(
            lambda row: is_stop_broken(row["現價"], row["最終ATR移動停利價"]),
            axis=1,
        ).sum()
    )
    elevated_margin_risk_count = int(report_df["融資風險"].map(is_elevated_margin_risk).sum())

    metric_columns = st.columns(6)
    metric_columns[0].metric("持股總市值", f"{market_value_total:,.2f}")
    metric_columns[1].metric("投入成本", f"{cost_total:,.2f}")
    metric_columns[2].metric("未實現損益", f"{unrealized_profit_total:,.2f}")
    metric_columns[3].metric("未實現損益率", f"{unrealized_profit_rate_total:.2f}%")
    metric_columns[4].metric("跌破移動停利股票數", f"{stop_break_count}")
    metric_columns[5].metric("融資風險升高股票數", f"{elevated_margin_risk_count}")

    st.subheader("完整報表")
    filter_columns = st.columns([2, 2, 2])
    search_text = filter_columns[0].text_input("搜尋股票代號或名稱", placeholder="例如 2382 或 廣達")
    risk_options = sorted(report_df["風險狀態"].dropna().astype(str).unique().tolist())
    selected_risk_status = filter_columns[1].multiselect("篩選風險狀態", risk_options)
    margin_risk_options = sorted(report_df["融資風險"].dropna().astype(str).unique().tolist())
    selected_margin_risk = filter_columns[2].multiselect("篩選融資風險", margin_risk_options)

    filtered_df = report_df.copy()
    if search_text:
        search_lower = search_text.strip().lower()
        filtered_df = filtered_df[
            filtered_df["股票代號"].astype(str).str.lower().str.contains(search_lower)
            | filtered_df["股票名稱"].astype(str).str.lower().str.contains(search_lower)
        ]

    if selected_risk_status:
        filtered_df = filtered_df[filtered_df["風險狀態"].isin(selected_risk_status)]

    if selected_margin_risk:
        filtered_df = filtered_df[filtered_df["融資風險"].isin(selected_margin_risk)]

    st.dataframe(style_report_dataframe(filtered_df), width="stretch", hide_index=True, height=560)

    attention_df = report_df[
        report_df["風險狀態"].isin(
            [
                "跌破ATR移動停利，建議分批停利",
                "低於成本且跌破ATR防守，建議減碼或停損",
                RISK_DOWNLOAD_FAILURE,
                RISK_DATA_INSUFFICIENT,
            ]
        )
        | report_df["融資風險"].map(is_elevated_margin_risk)
    ][["股票代號", "股票名稱", "風險狀態", "操作建議", "融資風險", "距離停利百分比"]]

    st.subheader("需要注意的股票清單")
    if attention_df.empty:
        st.success("目前沒有額外需要注意的股票。")
    else:
        st.dataframe(style_report_dataframe(attention_df), width="stretch", hide_index=True)


def render_download_buttons(
    report_df: pd.DataFrame,
    stop_history_df: pd.DataFrame,
    excel_error: str,
    key_prefix: str,
) -> None:
    """在主畫面與側邊欄共用下載按鈕內容。"""
    report_csv_bytes = create_csv_bytes(report_df)
    stop_history_csv_bytes = create_stop_history_csv_bytes(stop_history_df)

    st.download_button(
        "下載 atr_report.csv",
        data=report_csv_bytes,
        file_name="atr_report.csv",
        mime="text/csv",
        width="stretch",
        key=f"{key_prefix}-report-csv",
    )
    st.download_button(
        "下載最新 stop_history.csv",
        data=stop_history_csv_bytes,
        file_name="updated_stop_history.csv",
        mime="text/csv",
        width="stretch",
        key=f"{key_prefix}-stop-history-csv",
    )

    if excel_error:
        st.error(f"Excel 產生失敗：{excel_error}")
    else:
        st.download_button(
            "下載 atr_report.xlsx",
            data=st.session_state.get("report_excel_bytes", b""),
            file_name="atr_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
            key=f"{key_prefix}-report-xlsx",
        )


# 17. 串接側邊欄、計算流程、下載功能與主畫面渲染。
def main() -> None:
    """Streamlit 入口函式，負責整體頁面流程與互動控制。"""
    st.set_page_config(page_title="台股 ATR 移動停利系統", layout="wide")
    st.title("台股 ATR 移動停利與融資風險追蹤系統")
    st.caption(
        "透過 ATR、最近高點與融資籌碼變化，快速檢查持股的移動停利位置、潛在風險與操作建議。"
    )

    for key, default_value in {
        "report_df": pd.DataFrame(columns=REPORT_COLUMNS),
        "updated_stop_history_df": make_empty_stop_history_df(),
        "processing_messages": [],
        "report_excel_bytes": b"",
        "excel_error": "",
        "ocr_portfolio_df": make_empty_ocr_portfolio_df(),
        "ocr_raw_lines": [],
        "ocr_error": "",
    }.items():
        if key not in st.session_state:
            st.session_state[key] = default_value

    sample_portfolio_df = load_sample_portfolio()
    sample_portfolio_bytes = load_local_csv_bytes(PORTFOLIO_SAMPLE_PATH)
    sample_margin_bytes = load_local_csv_bytes(MARGIN_SAMPLE_PATH)
    sample_stop_history_bytes = load_local_csv_bytes(STOP_HISTORY_SAMPLE_PATH)

    with st.sidebar:
        st.header("資料輸入")
        portfolio_file = st.file_uploader("上傳 portfolio.csv", type=["csv"])
        portfolio_image_files = st.file_uploader(
            "上傳持股截圖（可多張）",
            type=PORTFOLIO_IMAGE_TYPES,
            accept_multiple_files=True,
            help="若券商系統無法匯出 CSV，可上傳持股截圖，系統會先做 OCR 草稿，再由你確認後計算。",
        )
        margin_file = st.file_uploader("上傳 margin.csv（選填，可覆蓋官方融資資料）", type=["csv"])
        stop_history_file = st.file_uploader("上傳 stop_history.csv（選填）", type=["csv"])

        st.caption("截圖請盡量裁成持股表格畫面，最好同時包含股票代號、名稱、股數與成本價。")
        ocr_default_category = st.selectbox(
            "截圖辨識預設 category",
            options=list(DEFAULT_ATR_MULTIPLIERS.keys()),
            index=list(DEFAULT_ATR_MULTIPLIERS.keys()).index("normal"),
        )
        share_unit_option = st.selectbox(
            "截圖中的持股數量單位",
            options=["股", "張（1張=1000股）"],
            index=0,
        )
        ocr_parse_button = st.button("辨識持股截圖", width="stretch")

        st.subheader("參數設定")
        atr_period = st.number_input("ATR 週期", min_value=2, max_value=60, value=5, step=1)
        high_period = st.number_input("最近高點期間", min_value=2, max_value=120, value=20, step=1)

        st.subheader("ATR 倍數設定")
        multiplier_settings = {
            "ETF": st.number_input("ETF 倍數", min_value=0.5, max_value=10.0, value=2.0, step=0.1),
            "financial": st.number_input("financial 倍數", min_value=0.5, max_value=10.0, value=2.0, step=0.1),
            "low_volatility": st.number_input(
                "low_volatility 倍數", min_value=0.5, max_value=10.0, value=2.0, step=0.1
            ),
            "normal": st.number_input("normal 倍數", min_value=0.5, max_value=10.0, value=2.0, step=0.1),
            "high_volatility": st.number_input(
                "high_volatility 倍數", min_value=0.5, max_value=10.0, value=2.5, step=0.1
            ),
        }

        calculate_button = st.button("開始計算", type="primary", width="stretch")

        st.subheader("範例檔下載")
        st.download_button(
            "下載 portfolio_sample.csv",
            data=sample_portfolio_bytes,
            file_name="portfolio_sample.csv",
            mime="text/csv",
            width="stretch",
            key="sidebar-sample-portfolio",
        )
        st.download_button(
            "下載 margin_sample.csv",
            data=sample_margin_bytes,
            file_name="margin_sample.csv",
            mime="text/csv",
            width="stretch",
            key="sidebar-sample-margin",
        )
        st.download_button(
            "下載 stop_history_sample.csv",
            data=sample_stop_history_bytes,
            file_name="stop_history_sample.csv",
            mime="text/csv",
            width="stretch",
            key="sidebar-sample-stop-history",
        )

        if not st.session_state["report_df"].empty:
            st.subheader("下載區")
            render_download_buttons(
                st.session_state["report_df"],
                st.session_state["updated_stop_history_df"],
                st.session_state["excel_error"],
                "sidebar",
            )

    if ocr_parse_button:
        if not portfolio_image_files:
            st.warning("尚未上傳持股截圖，請先選擇一張或多張圖片。")
        else:
            share_unit_multiplier = 1000 if share_unit_option.startswith("張") else 1
            try:
                ocr_portfolio_df = extract_portfolio_candidates_from_images(
                    portfolio_image_files,
                    default_category=ocr_default_category,
                    share_unit_multiplier=share_unit_multiplier,
                )
                st.session_state["ocr_portfolio_df"] = ocr_portfolio_df
                st.session_state["ocr_raw_lines"] = ocr_portfolio_df.attrs.get("raw_lines", [])
                st.session_state["ocr_error"] = ""
                if ocr_portfolio_df.empty:
                    st.warning("沒有從截圖辨識到可用的持股資料，請改上傳更清楚的截圖或手動使用 CSV。")
                else:
                    st.success("已完成持股截圖辨識，請先在下方草稿表確認與修正資料。")
            except Exception as exc:
                st.session_state["ocr_portfolio_df"] = make_empty_ocr_portfolio_df()
                st.session_state["ocr_raw_lines"] = []
                st.session_state["ocr_error"] = format_ocr_error_message(exc)
                st.error(f"截圖辨識失敗：{st.session_state['ocr_error']}")

    ocr_editor_df = None
    if st.session_state["ocr_error"]:
        st.error(f"OCR 模組或圖片處理發生問題：{st.session_state['ocr_error']}")

    if not st.session_state["ocr_portfolio_df"].empty:
        st.subheader("截圖辨識持股草稿")
        st.caption("請確認股票代號、股數、成本價與 category；若辨識不完整，可直接在下方表格手動修正後再按開始計算。")
        ocr_editor_df = st.data_editor(
            st.session_state["ocr_portfolio_df"],
            width="stretch",
            hide_index=True,
            num_rows="dynamic",
            column_config={
                "symbol": st.column_config.TextColumn("股票代號"),
                "name": st.column_config.TextColumn("股票名稱"),
                "shares": st.column_config.NumberColumn("股數", step=1, format="%.0f"),
                "cost": st.column_config.NumberColumn("成本價", step=0.01, format="%.2f"),
                "category": st.column_config.SelectboxColumn(
                    "category",
                    options=list(DEFAULT_ATR_MULTIPLIERS.keys()),
                    default=ocr_default_category,
                ),
                "image_name": st.column_config.TextColumn("來源截圖", disabled=True),
                "ocr_text": st.column_config.TextColumn("原始辨識文字", disabled=True),
            },
            key="ocr-portfolio-editor",
        )

        if st.session_state["ocr_raw_lines"]:
            with st.expander("查看 OCR 原始辨識文字"):
                for raw_line in st.session_state["ocr_raw_lines"]:
                    st.write(f"- {raw_line}")

    if calculate_button:
        portfolio_df = pd.DataFrame()
        portfolio_source = "csv"

        if portfolio_file is not None:
            try:
                portfolio_df = load_uploaded_csv(portfolio_file)
            except ValueError as exc:
                st.error(f"portfolio.csv 讀取失敗：{exc}")
                portfolio_df = pd.DataFrame()

            portfolio_df = finalize_portfolio_input(portfolio_df)
            missing_columns = validate_required_columns(portfolio_df, PORTFOLIO_REQUIRED_COLUMNS)
            if missing_columns:
                st.error(f"portfolio.csv 缺少必要欄位：{', '.join(missing_columns)}")
                portfolio_df = pd.DataFrame()
        elif ocr_editor_df is not None:
            portfolio_source = "ocr"
            portfolio_df = finalize_portfolio_input(ocr_editor_df, default_category=ocr_default_category)
            incomplete_ocr_df = portfolio_df[
                portfolio_df["symbol"].eq("") | portfolio_df["shares"].isna() | portfolio_df["cost"].isna()
            ]
            if not incomplete_ocr_df.empty:
                st.error("截圖辨識後仍有缺少股票代號、股數或成本價的列，請先在草稿表補齊再計算。")
                portfolio_df = pd.DataFrame()
        else:
            st.warning("請上傳 portfolio.csv，或先上傳持股截圖並按『辨識持股截圖』產生草稿。")

        if not portfolio_df.empty:
            margin_df = load_margin_data(margin_file, reference_date=date.today())
            stop_history_df = load_stop_history(stop_history_file)

            if portfolio_source == "ocr":
                st.info("本次使用截圖辨識後的持股草稿進行計算。")

            if margin_df.attrs.get("validation_error"):
                st.warning(f"margin.csv 讀取失敗，將略過融資分析：{margin_df.attrs['validation_error']}")
            if margin_df.attrs.get("missing_columns"):
                st.warning(
                    "margin.csv 缺少欄位，將略過融資分析："
                    + ", ".join(margin_df.attrs["missing_columns"])
                )
            if margin_df.attrs.get("source") == "official":
                if margin_df.attrs.get("source_summary"):
                    st.info(margin_df.attrs["source_summary"])
                if margin_df.attrs.get("fetch_warnings") and not margin_df.attrs.get("fetch_error"):
                    st.warning("部分官方融資資料未取得：" + "；".join(margin_df.attrs["fetch_warnings"]))
                if margin_df.attrs.get("fetch_error"):
                    st.warning(f"官方融資資料自動抓取失敗：{margin_df.attrs['fetch_error']}")

            if stop_history_df.attrs.get("validation_error"):
                st.warning(f"stop_history.csv 讀取失敗，將從空白歷史開始：{stop_history_df.attrs['validation_error']}")
            if stop_history_df.attrs.get("missing_columns"):
                st.warning(
                    "stop_history.csv 缺少欄位，將從空白歷史開始："
                    + ", ".join(stop_history_df.attrs["missing_columns"])
                )

            settings = {
                "atr_period": int(atr_period),
                "high_period": int(high_period),
                "multiplier_settings": multiplier_settings,
            }

            with st.spinner("正在下載股價、計算 ATR 與整理報表..."):
                report_df = build_report(portfolio_df, margin_df, stop_history_df, settings)
                updated_stop_history_df = build_updated_stop_history(report_df, stop_history_df)

            st.session_state["report_df"] = report_df
            st.session_state["updated_stop_history_df"] = updated_stop_history_df
            st.session_state["processing_messages"] = report_df.attrs.get("processing_messages", [])

            try:
                st.session_state["report_excel_bytes"] = create_excel_bytes(report_df)
                st.session_state["excel_error"] = ""
            except Exception as exc:
                st.session_state["report_excel_bytes"] = b""
                st.session_state["excel_error"] = str(exc)

    if portfolio_file is None and st.session_state["ocr_portfolio_df"].empty and st.session_state["report_df"].empty:
        st.info("請先上傳 portfolio.csv，或改上傳持股截圖進行 OCR 辨識，系統才會開始計算 ATR、移動停利與融資風險。")
        st.download_button(
            "下載 portfolio_sample.csv",
            data=sample_portfolio_bytes,
            file_name="portfolio_sample.csv",
            mime="text/csv",
            key="main-sample-portfolio",
        )
        st.subheader("portfolio_sample.csv 預覽")
        st.dataframe(sample_portfolio_df, width="stretch", hide_index=True)

    if not st.session_state["report_df"].empty:
        render_dashboard(st.session_state["report_df"])

        if st.session_state["processing_messages"]:
            with st.expander("處理訊息與警示"):
                for message in st.session_state["processing_messages"]:
                    st.write(f"- {message}")

        st.subheader("下載區")
        render_download_buttons(
            st.session_state["report_df"],
            st.session_state["updated_stop_history_df"],
            st.session_state["excel_error"],
            "main",
        )


if __name__ == "__main__":
    main()